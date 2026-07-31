#!/usr/bin/env python3
"""
Convert CRM In Transit Report Excel to in_transit_index.json
Based on Pakistan training document approach:
- Filter: materialStatusName == 'Good'
- Group by material code + toWarehouseName
- Sum quantities
"""
import json
import sys
import openpyxl

def build_in_transit_index(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}
    
    material_col = col_map.get('materialCode')
    warehouse_col = col_map.get('toWarehouseName')
    status_col = col_map.get('materialStatusName')
    qty_col = col_map.get('quantity')
    
    if None in [material_col, warehouse_col, status_col, qty_col]:
        print(f"ERROR: Missing columns. Found: {headers}")
        sys.exit(1)
    
    in_transit = {}
    good_count = 0
    total_count = 0
    skipped = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        total_count += 1
        
        status = str(row[status_col]).strip() if row[status_col] else ''
        if status != 'Good':
            skipped += 1
            continue
        
        material = str(row[material_col]).strip() if row[material_col] else ''
        warehouse = str(row[warehouse_col]).strip() if row[warehouse_col] else ''
        qty = row[qty_col] if row[qty_col] else 0
        
        if not material or not warehouse:
            continue
        
        qty = int(qty) if qty else 0
        
        if material not in in_transit:
            in_transit[material] = {}
        
        in_transit[material][warehouse] = in_transit[material].get(warehouse, 0) + qty
        good_count += 1
    
    wb.close()
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(in_transit, f, ensure_ascii=False, indent=2)
    
    print(f"Total rows: {total_count}")
    print(f"Good status: {good_count}")
    print(f"Skipped: {skipped}")
    print(f"Unique materials: {len(in_transit)}")
    print(f"Output: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: build_in_transit_index.py <input.xlsx> <output.json>")
        sys.exit(1)
    build_in_transit_index(sys.argv[1], sys.argv[2])