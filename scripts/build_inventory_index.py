#!/usr/bin/env python3
"""
Convert CRM Inventory Overview Excel to inventory_index.json
Based on Pakistan training document approach:
- Filter: status == 'Good'
- Use 'Available' column for quantity
- Use 'Warehouse Name' column for warehouse key
- Include '_desc' field with material description
"""
import json
import sys
import openpyxl

def build_inventory_index(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Find column indices
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}
    
    warehouse_col = col_map.get('Warehouse Name')
    material_col = col_map.get('Assets/Material Code')
    desc_col = col_map.get('Assets/Material Desc')
    status_col = col_map.get('Status')
    available_col = col_map.get('Available')
    
    if None in [warehouse_col, material_col, status_col, available_col]:
        print(f"ERROR: Missing columns. Found: {headers}")
        sys.exit(1)
    
    inventory = {}
    good_count = 0
    total_count = 0
    skipped = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        total_count += 1
        
        status = str(row[status_col]).strip() if row[status_col] else ''
        if status != 'Good':
            skipped += 1
            continue
        
        warehouse = str(row[warehouse_col]).strip() if row[warehouse_col] else ''
        material = str(row[material_col]).strip() if row[material_col] else ''
        available = row[available_col] if row[available_col] else 0
        desc = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else ''
        
        if not material or not warehouse:
            continue
        
        available = int(available) if available else 0
        if available < 0:
            available = 0
        
        if material not in inventory:
            inventory[material] = {}
        
        if warehouse not in inventory[material]:
            inventory[material][warehouse] = {"Good": 0}
        
        inventory[material][warehouse]["Good"] += available
        
        if desc and '_desc' not in inventory[material]:
            inventory[material]['_desc'] = desc
        
        good_count += 1
    
    wb.close()
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(inventory, f, ensure_ascii=False, indent=2)
    
    print(f"Total rows: {total_count}")
    print(f"Good status: {good_count}")
    print(f"Skipped (non-Good): {skipped}")
    print(f"Unique materials: {len(inventory)}")
    print(f"Output: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: build_inventory_index.py <input.xlsx> <output.json>")
        sys.exit(1)
    build_inventory_index(sys.argv[1], sys.argv[2])